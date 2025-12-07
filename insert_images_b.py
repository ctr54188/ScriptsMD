import os
import sys
import warnings
from urllib3.exceptions import InsecureRequestWarning
warnings.filterwarnings("ignore", category=InsecureRequestWarning)

import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image as PILImage
from tqdm import tqdm  # 导入tqdm库

# === 必须放在最前面！===
if sys.platform == "darwin":  # 仅 macOS
    # 禁用 multiprocessing 的 resource tracker
    os.environ["OBJC_DISABLE_INITIALIZE_FORK_SAFETY"] = "YES"
    # 关键：防止 resource_tracker 启动
    import multiprocessing
    multiprocessing.set_start_method("fork", force=True)
    # 或者更彻底地禁用（PyInstaller 推荐）
    from multiprocessing import util
    util._finalizer_registry.clear()
# ======================

def process_excel(input_path):
    if not os.path.exists(input_path):
        print(f"❌ 文件不存在: {input_path}")
        return False

    output_path = os.path.splitext(input_path)[0] + "_含图片版.xlsx"
    print(f"📂 处理文件: {input_path}")

    wb = load_workbook(input_path)
    ws = wb.active

    TARGET_COL_WIDTH = 34.75
    TARGET_ROW_HEIGHT = 141
    total = 0

    # 使用 tqdm 创建进度条
    for row in tqdm(range(1, ws.max_row + 1), desc="进度", ncols=80):  # 这里加入了tqdm进度条
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            url = cell.value

            if url and isinstance(url, str) and url.strip().startswith(("http://", "https://")):
                url = url.strip()
                try:
                    resp = requests.get(url, timeout=15, verify=False)
                    resp.raise_for_status()

                    img_data = BytesIO(resp.content)
                    pil_img = PILImage.open(img_data)
                    if pil_img.mode != "RGB":
                        pil_img = pil_img.convert("RGB")

                    # 换算像素
                    target_w = int(TARGET_COL_WIDTH * 7.2)
                    target_h = int(TARGET_ROW_HEIGHT * 1.33)
                    pil_img = pil_img.resize((target_w, target_h), PILImage.LANCZOS)

                    img_bytes = BytesIO()
                    pil_img.save(img_bytes, "JPEG", quality=90)
                    img_bytes.seek(0)

                    xl_img = XLImage(img_bytes)
                    xl_img.width = target_w
                    xl_img.height = target_h

                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = TARGET_COL_WIDTH
                    ws.row_dimensions[row].height = TARGET_ROW_HEIGHT

                    ws.add_image(xl_img, cell.coordinate)
                    cell.value = None
                    total += 1

                except Exception as e:
                    print(f"❌ 行{row}列{col}失败: {str(e)[:100]}")

    wb.save(output_path)
    print(f"\n✅ 完成！结果: {output_path}")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("💡 用法: 将 Excel 文件拖到本程序上，或命令行运行:")
        print(f"      {os.path.basename(sys.executable)} <文件路径>")
        input("\n按回车退出...")
        sys.exit(1)

    input_file = sys.argv[1]
    if not input_file.endswith(('.xlsx', '.xls')):
        print("⚠️ 仅支持 .xlsx 文件")
        input("按回车退出...")
        sys.exit(1)

    try:
        process_excel(input_file)
        input("\n✅ 处理完成！按回车退出...")
    except Exception as e:
        print(f"💥 严重错误: {e}")
        input("按回车退出...")
        sys.exit(1)